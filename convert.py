# convert.py

import os
import pandas as pd
import datetime
from openpyxl import load_workbook
from win32com.client import Dispatch
from shutil import copyfile

import yg_fl
import jjr_fl
import sb_fl
import gjj_fl


class Convert:
    pass


def cshfl(zdr_bm, yyb_bm):  # 初始化分录
    today = datetime.date.today().strftime("%Y-%m-%d")
    kjfl = [  # 会计分录
        None,  # 0  A列
        yyb_bm + '-0002',  # 1  B列，核算账簿
        '01',  # 2  C列，凭证类别编码
        zdr_bm,  # 3  D列，制单人编码
        today,  # 4  E列，制单日期
        None,  # 5  F列，摘要
        None,  # 6  G列，科目编码
        '人民币',  # 7  H列，币种
        None,  # 8  I列，原币借方金额
        None,  # 9  J列，本币借方金额
        yyb_bm,  # 10 K列，业务单元编码
        None,  # 11 L列，原币贷方金额
        None,  # 12 M列，本币贷方金额
        today,  # 13 N列，业务日期
        '1',  # 14 O列，组织本币汇率
        None,  # 15 P列，辅助核算1
        None,  # 16 Q列，辅助核算2
        None,  # 17 R列，辅助核算3
    ]
    return kjfl


def SInfo():
    global SInfo_df
    SInfo_xlFile = "docs\\结算信息.xlsx"    # 结算信息excel文件
    SInfo_df = pd.read_excel(SInfo_xlFile, index_col=2 - 1, skiprows=1 - 1)
    return SInfo_df


SInfo_df = SInfo()


def create_dir_not_exist(path):
    if not os.path.exists(path):
        os.mkdir(path)


create_dir_not_exist("output\\")


def xls_win32save(xlfile):                 # 用win32com组件将out_xls打开再保存，这样用友才能识别其中内容。
    xl = Dispatch("Excel.Application")
    # 后台运行，不显示，不警告
    xl.Visible = False
    xl.DisplayAlerts = False

    xlfile = os.path.abspath(xlfile)
    wb = xl.Workbooks.Open(xlfile)          # win32不认识相对路径，故需上一句转换为绝对路径。

    # wb.SaveAs(xlfile)
    wb.Save()
    # 关闭表格和excel对象
    wb.Close()
    xl.Quit()


def out_xls(yyb_bm, pz):
    out_xlfile = "output\\" + yyb_bm + "_" + pz + ".xlsx"
    copyfile("docs\\template.xlsx", out_xlfile)
    out_wb = load_workbook(out_xlfile)
    out_ws = out_wb.active
    return out_xlfile, out_wb, out_ws


def convert_yg(zdr_bm, yyb_bm, jbb_bm, in_df):
    global amount_A, amount_B, amount_D, amount_L
    out_xlfile, out_wb, out_ws = out_xls(yyb_bm, "yg")

    # 获取字典的所有键
    kmdm_col = in_df.columns
    kmdm_AB = yg_fl.dict_AB.keys()           # 字典dict键key列表
    kmdm_Z = yg_fl.dict.keys()

    xj_list = [  # 小计列表
        "22110103",  # 补贴小计
        "221102",  # 福利小计
        "22110104",  # 提成小计
    ]

    bt_list = [  # 补贴列表
        "66011503",  # 过节费
        "66011504",  # 交通补贴
        "66011505",  # 伙食补贴
        "66011506",  # 通讯补贴
        # "66011507",  # 辞退福利
        "66011510",  # 劳保补贴
        "66011519",  # 其他补贴
    ]

    flf_list = [k for k in kmdm_AB if k[0:6] == "660116"]   # 福利费列表       在kmdm_AB列表中提取福利费科目代码列表
    tc_list = [k for k in kmdm_AB if k[0:8] == "66011508"]  # 提成列表         在kmdm_AB列表中提取提成支出科目代码列表

    amt_a_bt = 0  # A类人员补贴、福利、提成小计金额
    amt_a_fl = 0
    amt_a_tc = 0

    amt_b_bt = 0  # B类人员补贴、福利、提成小计金额
    amt_b_fl = 0
    amt_b_tc = 0

    amt_d_bt = 0  # D类人员补贴、福利、提成小计金额
    amt_d_fl = 0
    amt_d_tc = 0

    amt_l_bt = 0  # L类人员补贴、福利、提成小计金额
    amt_l_fl = 0
    amt_l_tc = 0

    if 'A小计' in in_df.index:
        # A类人员（员工）-01
        for k in kmdm_AB:                                        # 做账循序科目代码列表for循环
            fl_list = cshfl(zdr_bm, yyb_bm)                      # 初始化分录各字段list列表
            if (k in kmdm_col) or (k in xj_list):                # 元素是否在表头col(科目代码)或小计列表中
                if k in kmdm_col:
                    amount_A = in_df[k]['A小计']                  # 取“A小计”行数据
                    if (amount_A != 0) and (not (pd.isnull(amount_A))):
                        amount_A = round(in_df[k]['A小计'], 2)    # 对合计数据四舍五入，并转为字符型。
                        if k in bt_list:
                            amt_a_bt += amount_A
                        if k in flf_list:
                            amt_a_fl += amount_A
                        if k in tc_list:
                            amt_a_tc += amount_A
                        if k == '22110103':
                            amount_A = amt_a_bt
                        if k == '221102':
                            amount_A = amt_a_fl
                        if k == '22110104':
                            amount_A = amt_a_tc
                        yg_fl.switcher(yg_fl.dict_AB, k, fl_list, amount_A, '01:员工', jbb_bm, in_df, out_ws)

                if (k not in kmdm_col) and (k in xj_list):
                    if k == '22110103':
                        amount_A = amt_a_bt
                    if k == '221102':
                        amount_A = amt_a_fl
                    if k == '22110104':
                        amount_A = amt_a_tc
                    if amount_A != 0:
                        yg_fl.switcher(yg_fl.dict_AB, k, fl_list, amount_A, '01:员工', jbb_bm, in_df, out_ws)

    if 'B小计' in in_df.index:
        # B类人员（全日制营销人员）-05
        for k in kmdm_AB:                                        # 做账循序科目代码列表for循环
            fl_list = cshfl(zdr_bm, yyb_bm)                      # 初始化分录各字段list列表
            if (k in kmdm_col) or (k in xj_list):                # 元素是否在表头col(科目代码)或小计列表中
                if k in kmdm_col:
                    amount_B = in_df[k]['B小计']                  # 取“B小计”行数据
                    if (amount_B != 0) and (not (pd.isnull(amount_B))):     # 当有多个“B小计”时，amount_B就不是一个数值
                        amount_B = round(in_df[k]['B小计'], 2)    # 对合计数据四舍五入，并转为字符型。
                        if k in bt_list:
                            amt_b_bt += amount_B
                        if k in flf_list:
                            amt_b_fl += amount_B
                        if k in tc_list:
                            amt_b_tc += amount_B
                        if k == '22110103':
                            amount_B = amt_b_bt
                        if k == '221102':
                            amount_B = amt_b_fl
                        if k == '22110104':
                            amount_B = amt_b_tc
                        yg_fl.switcher(yg_fl.dict_AB, k, fl_list, amount_B, '05:营销人员', jbb_bm, in_df, out_ws)

                if (k not in kmdm_col) and (k in xj_list):
                    if k == '22110103':
                        amount_B = amt_b_bt
                    if k == '221102':
                        amount_B = amt_b_fl
                    if k == '22110104':
                        amount_B = amt_b_tc
                    if amount_B != 0:
                        yg_fl.switcher(yg_fl.dict_AB, k, fl_list, amount_B, '05:营销人员', jbb_bm, in_df, out_ws)

    if 'D小计' in in_df.index:
        # D类人员（实习生）-08
        for k in kmdm_AB:                                        # 做账循序科目代码列表for循环
            fl_list = cshfl(zdr_bm, yyb_bm)                      # 初始化分录各字段list列表
            if (k in kmdm_col) or (k in xj_list):                # 元素是否在表头col(科目代码)或小计列表中
                if k in kmdm_col:
                    amount_D = in_df[k]['D小计']                  # 取“D小计”行数据
                    if (amount_D != 0) and (not (pd.isnull(amount_D))):
                        amount_D = round(in_df[k]['D小计'], 2)    # 对合计数据四舍五入，并转为字符型。
                        if k in bt_list:
                            amt_d_bt += amount_D
                        if k in flf_list:
                            amt_d_fl += amount_D
                        if k in tc_list:
                            amt_d_tc += amount_D
                        if k == '22110103':
                            amount_D = amt_d_bt
                        if k == '221102':
                            amount_D = amt_d_fl
                        if k == '22110104':
                            amount_D = amt_d_tc
                        yg_fl.switcher(yg_fl.dict_D, k, fl_list, amount_D, '08:实习生', jbb_bm, in_df, out_ws)

                if (k not in kmdm_col) and (k in xj_list):
                    if k == '22110103':
                        amount_D = amt_d_bt
                    if k == '221102':
                        amount_D = amt_d_fl
                    if k == '22110104':
                        amount_D = amt_d_tc
                    if amount_D != 0:
                        yg_fl.switcher(yg_fl.dict_D, k, fl_list, amount_D, '08:实习生', jbb_bm, in_df, out_ws)

    if 'L小计' in in_df.index:
        # L类人员（劳务）-09
        for k in kmdm_AB:                                        # 做账循序科目代码列表for循环
            fl_list = cshfl(zdr_bm, yyb_bm)                      # 初始化分录各字段list列表
            if (k in kmdm_col) or (k in xj_list):                # 元素是否在表头col(科目代码)或小计列表中
                if k in kmdm_col:
                    amount_L = in_df[k]['L小计']                  # 取“L小计”行数据
                    if (amount_L != 0) and (not (pd.isnull(amount_L))):
                        amount_L = round(in_df[k]['L小计'], 2)    # 对合计数据四舍五入，并转为字符型。
                        if k in bt_list:
                            amt_l_bt += amount_L
                        if k in flf_list:
                            amt_l_fl += amount_L
                        if k in tc_list:
                            amt_l_tc += amount_L
                        if k == '22110103':
                            amount_L = amt_l_bt
                        if k == '221102':
                            amount_L = amt_l_fl
                        if k == '22110104':
                            amount_L = amt_l_tc
                        yg_fl.switcher(yg_fl.dict_L, k, fl_list, amount_L, '09:劳务', jbb_bm, in_df, out_ws)

                if (k not in kmdm_col) and (k in xj_list):
                    if k == '22110103':
                        amount_L = amt_l_bt
                    if k == '221102':
                        amount_L = amt_l_fl
                    if k == '22110104':
                        amount_L = amt_l_tc
                    if amount_L != 0:
                        yg_fl.switcher(yg_fl.dict_L, k, fl_list, amount_L, '09:劳务', jbb_bm, in_df, out_ws)

    if '负责人小计' in in_df.index:
        pass

    # 全部人员合计
    for k in kmdm_Z:                                         # 做账循序科目代码列表for循环
        fl_list = cshfl(zdr_bm, yyb_bm)                      # 初始化分录各字段list列表
        if k in kmdm_col:                                    # 做账循序科目代码列表元素是否在表头col(科目代码)中
            amount = in_df[k]['合计']                         # 取“合计”行数据
            if (amount != 0) and (not (pd.isnull(amount))):
                amount = round(in_df[k]['合计'], 2)           # 对合计数据四舍五入，并转为字符型。
                yg_fl.switcher(yg_fl.dict, k, fl_list, amount, '01:员工', jbb_bm, in_df, out_ws)

    # 支付结算分录
    fl_list = cshfl(zdr_bm, yyb_bm)  # 第四次、初始化分录各字段list列表
    amt = round(in_df['1001']['合计'], 2)  # 对合计数据四舍五入，并转为字符型。
    yg_fl.km1001(SInfo_df, fl_list, amt, jbb_bm, in_df, out_ws)

    out_wb.save(filename=out_xlfile)
    out_wb.close()
    xls_win32save(out_xlfile)

    # # https://stackoverflow.com/questions/56960564/how-do-i-format-an-entire-column-or-cells-i-can-iterate-thruas-text-format-usi
    # for row in out_ws.iter_rows(min_row=2):                      # min_row=2  从第2行开始，排除第1行。
    #     for cell in row:
    #         # only relevant column and without header
    #         # if cell.column_letter == 'D' and cell.row > 1:
    #         out_ws[cell.coordinate].number_format = '@'          # @ 文本格式
    #         # out_ws['A1'].number_format = 'General'             # General 格式


def convert_jjr(zdr_bm, yyb_bm, jbb_bm, in_df):
    out_xlfile, out_wb, out_ws = out_xls(yyb_bm, "jjr")

    for k in in_df.columns:  # col是列字段名(科目代码)
        fl_list = cshfl(zdr_bm, yyb_bm)  # 初始化分录各字段list列表
        amount = in_df[k]['合计']  # 从左至右取每列/字段的最后一行的合计数据
        if (amount != 0) and (not (pd.isnull(amount))):
            amount = round(in_df[k]['合计'], 2)  # 对合计数据四舍五入，并转为字符型。
            jjr_fl.switcher(jjr_fl.case, k, fl_list, amount, jbb_bm, in_df, out_ws)

    # 支付结算分录
    fl_list = cshfl(zdr_bm, yyb_bm)  # 第四次、初始化分录各字段list列表
    amt = round(in_df['1001']['合计'], 2)  # 对合计数据四舍五入，并转为字符型。
    jjr_fl.km1001(SInfo_df, fl_list, amt, jbb_bm, in_df, out_ws)

    out_wb.save(filename=out_xlfile)
    out_wb.close()
    xls_win32save(out_xlfile)


def convert_sb(zdr_bm, yyb_bm, jbb_bm, in_df, sf_df, dz_df):
    out_xlfile, out_wb, out_ws = out_xls(yyb_bm, "sb")

    # 一、 单位部分社保分录
    for k in in_df.columns:  # col是列字段名(科目代码)
        fl_list = cshfl(zdr_bm, yyb_bm)  # 第一次、初始化分录各字段list列表
        amount = in_df[k]['成本数据']  # 从左至右取每列/字段的最后一行的合计数据
        if (amount != 0) and (not (pd.isnull(amount))):
            amount = round(in_df[k]['成本数据'], 2)  # 对合计数据四舍五入，并转为字符型。
            sb_fl.switcher(sb_fl.case_1, k, fl_list, amount, jbb_bm, in_df, dz_df, out_ws)

    # 二、 个人部分社保分录
    for k in in_df.columns:  # col是列字段名(科目代码)
        fl_list = cshfl(zdr_bm, yyb_bm)  # 第二次、初始化分录各字段list列表
        amount = in_df[k]['成本数据']  # 从左至右取每列/字段的最后一行的合计数据
        if (amount != 0) and (not (pd.isnull(amount))):
            amount = round(in_df[k]['成本数据'], 2)  # 对合计数据四舍五入，并转为字符型。
            sb_fl.switcher(sb_fl.case_2, k, fl_list, amount, jbb_bm, in_df, dz_df, out_ws)

    # 三、代垫总部分录
    fl_list = cshfl(zdr_bm, yyb_bm)  # 第三次、初始化分录各字段list列表
    sb_fl.km_dzgr(fl_list, dz_df, out_ws)

    # 四、应收应付分录
    fl_list = cshfl(zdr_bm, yyb_bm)  # 第三次、初始化分录各字段list列表
    sb_fl.km_sfsb(fl_list, sf_df, out_ws)

    # 五、银行付款这一笔分录
    fl_list = cshfl(zdr_bm, yyb_bm)  # 第四次、初始化分录各字段list列表
    sb_fl.km1001(SInfo_df, fl_list, in_df, out_ws)

    out_wb.save(filename=out_xlfile)
    out_wb.close()
    xls_win32save(out_xlfile)


def convert_gjj(zdr_bm, yyb_bm, jbb_bm, in_df, sf_df, dz_df):
    out_xlfile, out_wb, out_ws = out_xls(yyb_bm, "gjj")

    # 一、 单位、个人公积金分录
    for k in in_df.columns:  # col是列字段名(科目代码)
        fl_list = cshfl(zdr_bm, yyb_bm)  # 第一次、初始化分录各字段list列表
        amount = in_df[k]['成本数据']  # 从左至右取每列/字段的最后一行的合计数据
        if (amount != 0) and (not (pd.isnull(amount))):
            amount = round(in_df[k]['成本数据'], 2)  # 对合计数据四舍五入，并转为字符型。
            gjj_fl.switcher(gjj_fl.case, k, fl_list, amount, jbb_bm, in_df, dz_df, out_ws)

    # 二、代垫总部分录
    fl_list = cshfl(zdr_bm, yyb_bm)  # 第三次、初始化分录各字段list列表
    gjj_fl.km_dzgr(fl_list, dz_df, out_ws)

    # 三、应收应付分录
    fl_list = cshfl(zdr_bm, yyb_bm)  # 第二次、初始化分录各字段list列表
    gjj_fl.km_sfgjj(fl_list, sf_df, out_ws)

    # 四、银行付款这一笔分录
    fl_list = cshfl(zdr_bm, yyb_bm)  # 第三次、初始化分录各字段list列表
    gjj_fl.km1001(SInfo_df, fl_list, in_df, out_ws)

    out_wb.save(filename=out_xlfile)
    out_wb.close()
    xls_win32save(out_xlfile)


if __name__ == '__main__':
    pass
    # ZDR_bm = 'wms'
    # YYB_bm = '1245'
    # JBB_bm = '1101'
    # LBT_col = 4
    # HBT_idx = 4
    # Filename = ".\\docs\\yggz_2020年9月工资表明细-珠海营业部.xlsx"
    # convert_jjrgz(ZDR_bm, YYB_bm, JBB_bm, LBT_col, HBT_idx, Filename)
